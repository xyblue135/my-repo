import os
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook

def count_value_in_all_files(value):
    directory = os.getcwd()
    xlsx_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    csv_files = [f for f in os.listdir(directory) if f.endswith('.csv')]
    
    total_count = 0
    details = []

    # 遍历每个xlsx文件
    for xlsx_file in xlsx_files:
        file_path = os.path.join(directory, xlsx_file)
        workbook = load_workbook(file_path, read_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            count = 0

            # 遍历sheet中的每一行
            for row in sheet.iter_rows(values_only=True):
                if value in row:
                    count += 1

            if count > 0:
                details.append(f"文件 {xlsx_file} 中的工作表 {sheet_name} 共有 {count} 行包含值 {value}")
                total_count += count

    # 遍历每个csv文件
    for csv_file in csv_files:
        file_path = os.path.join(directory, csv_file)
        df = pd.read_csv(file_path)
        count = df.apply(lambda row: value in row.values, axis=1).sum()

        if count > 0:
            details.append(f"文件 {csv_file} 共有 {count} 行包含值 {value}")
            total_count += count

    result_message = f"总共有 {total_count} 行包含值 {value}"
    if details:
        details_message = "\n".join(details)
        messagebox.showinfo("详细结果", details_message)
    messagebox.showinfo("总计", result_message)

def on_search_click():
    value_to_find = value_entry.get()
    count_value_in_all_files(value_to_find)

app = tk.Tk()
app.title("查找值并统计行数")

tk.Label(app, text="请输入要查找的值:").pack(pady=5)
value_entry = tk.Entry(app, width=50)
value_entry.pack(pady=5)

tk.Button(app, text="开始查找", command=on_search_click).pack(pady=20)

app.mainloop()
