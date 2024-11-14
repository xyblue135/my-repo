import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def process_txt_to_excel(file_path):
    try:
        # 读取txt文件内容
        with open(file_path, 'r') as file:
            data_lines = file.readlines()

        # 处理数据，去除引号和管道符号，分割成列表
        data_rows = [line.strip().replace('"', '').split('|') for line in data_lines]

        # 检查是否有足够的数据行进行转换
        if len(data_rows) < 2:
            raise ValueError("文件内容不足以构成有效的表格数据")

        # 创建DataFrame
        df = pd.DataFrame(data_rows[1:], columns=data_rows[0])  # 假设第一行为列名

        # 构建输出文件名，将.txt替换为.xlsx
        excel_file_path = file_path.replace('.txt', '.xlsx')

        # 保存为Excel文件
        df.to_excel(excel_file_path, index=False)

        return df, os.path.basename(excel_file_path).replace('.xlsx', '')  # 返回DataFrame和表名

    except Exception as e:
        print(f"处理 {file_path} 时发生错误: {e}")
        return None, None

def format_excel_sheets(excel_file_path):
    wb = load_workbook(excel_file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        # 取消第一行字体加粗
        for cell in ws[1]:
            cell.font = Font(bold=False)  # 使用Font类创建非加粗字体
        
        # 设置单元格对齐方式
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
    
    wb.save(excel_file_path)

# 获取当前目录下所有.txt文件
txt_files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.txt')]

all_dfs = []
sheet_names = []

# 遍历并处理每个.txt文件
for txt_file in txt_files:
    df, sheet_name = process_txt_to_excel(txt_file)
    if df is not None:  # 确保转换成功才添加到列表
        all_dfs.append(df)
        sheet_names.append(sheet_name)
        print(f"{txt_file} 成功转换为 {sheet_name}.xlsx")
    else:
        print(f"{txt_file} 转换失败。")

if all_dfs:
    with pd.ExcelWriter('汇总.xlsx', engine='openpyxl') as writer:
        for df, sheet_name in zip(all_dfs, sheet_names):
            df.to_excel(writer, sheet_name=sheet_name, index=False)  # 直接使用txt文件名作为sheet名

    # 格式化Excel文件中的所有工作表
    format_excel_sheets('汇总.xlsx')
    print("所有数据已合并至汇总.xlsx，每个工作表命名对应原.txt文件名。")
else:
    print("没有数据可以合并，因为所有.txt文件都未能成功转换。")

print("处理完成。")
